import tkinter as tk
from tkinter import messagebox, ttk
from tkinter. ttk import Treeview
from tkinter.filedialog import asksaveasfilename
from datetime import datetime
import openpyxl
from database import get_db_connection

class LibraryStatisticsScreen:
    def __init__(self, root):
        self.root = root
        self.root.title("Thống kê sách")
        self.root.geometry("1100x800")
        self.root.configure(bg="#babfbb")
        self.conn = get_db_connection()
        self.cursor = self.conn.cursor()
        self.setup_ui()
        self.refresh_data()

    def setup_ui(self):
        # Cấu hình grid weight để responsive
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.lb_background = "#dcdad5"
        
        # Main frame với scrollbar
        main_container = ttk.Frame(self.root, style="Main.TFrame")
        main_container.grid(row=0, column=0, sticky="nsew")
        main_container.grid_rowconfigure(0, weight=1)
        main_container.grid_columnconfigure(0, weight=1)
        
        # Canvas và Scrollbar
        canvas = tk.Canvas(main_container)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        
        self.frame_statics = ttk.Frame(canvas)
        
        # Cấu hình canvas
        canvas.configure(yscrollcommand=scrollbar.set, background=self.lb_background)
        scrollbar.grid(row=0, column=1, sticky="ns")
        canvas.grid(row=0, column=0, sticky="nsew")
        canvas.create_window((0, 0), window=self.frame_statics, anchor="nw")
        
        # Cập nhật scroll region
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        self.frame_statics.bind("<Configure>", on_frame_configure)
        
        # Bind mousewheel
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)

        # Nút Trở về
        btn_back = ttk.Button(self.frame_statics, text="← Trở về", command=self.tro_ve)
        btn_back.grid(row=0, column=0, sticky="W", padx=10, pady=10)

        # Tiêu đề
        tk.Label(self.frame_statics, text="THỐNG KÊ THƯ VIỆN", font=("Arial", 16, "bold"), background=self.lb_background).grid(
            row=1, column=0, columnspan=4, pady=(0, 20), sticky="n")

        # Labels cho thống kê
        self.lbl_total_books = tk.Label(self.frame_statics, text="0", font=("Arial", 10), background=self.lb_background)
        self.lbl_total_readers = tk.Label(self.frame_statics, text="0", font=("Arial", 10), background=self.lb_background)
        self.lbl_remaining_books = tk.Label(self.frame_statics, text="0", font=("Arial", 10), background=self.lb_background)
        self.lbl_borrowed_books = tk.Label(self.frame_statics, text="0", font=("Arial", 10), background=self.lb_background)
        self.lbl_return_rate = tk.Label(self.frame_statics, text="0%", font=("Arial", 10), background=self.lb_background)
        self.lbl_most_borrowed_book_val = tk.Label(self.frame_statics, text="", font=("Arial", 10), background=self.lb_background)
        self.lbl_most_active_reader_val = tk.Label(self.frame_statics, text="", font=("Arial", 10), background=self.lb_background)
        self.lbl_books_not_returned_count = tk.Label(self.frame_statics, text="Số lượng: 0", font=("Arial", 10), background=self.lb_background)
        self.lbl_books_due_not_returned_count = tk.Label(self.frame_statics, text="Số lượng: 0", font=("Arial", 10), background=self.lb_background)

        # Các thống kê cơ bản
        tk.Label(self.frame_statics, text="Tổng số lượng sách:", font=("Arial", 10, "bold"), background=self.lb_background).grid(
            row=2, column=0, sticky="w", padx=10, pady=5)
        self.lbl_total_books.grid(row=2, column=1, sticky="w", padx=10)

        tk.Label(self.frame_statics, text="Tổng số độc giả:", font=("Arial", 10, "bold"), background=self.lb_background).grid(
            row=2, column=2, sticky="w", padx=10, pady=5)
        self.lbl_total_readers.grid(row=2, column=3, sticky="w", padx=10)

        tk.Label(self. frame_statics, text="Số sách đang được mượn:", font=("Arial", 10, "bold"), background=self.lb_background).grid(
            row=3, column=0, sticky="w", padx=10, pady=5)
        self.lbl_borrowed_books.grid(row=3, column=1, sticky="w", padx=10)

        tk.Label(self.frame_statics, text="Số sách còn lại trong kho:", font=("Arial", 10, "bold"), background=self.lb_background).grid(
            row=3, column=2, sticky="w", padx=10, pady=5)
        self.lbl_remaining_books.grid(row=3, column=3, sticky="w", padx=10)

        tk. Label(self.frame_statics, text="Tỉ lệ trả sách đúng hạn:", font=("Arial", 10, "bold"), background=self.lb_background).grid(
            row=4, column=0, sticky="w", padx=10, pady=5)
        self.lbl_return_rate.grid(row=4, column=1, sticky="w", padx=10)

        tk.Label(self.frame_statics, text="Sách được mượn nhiều nhất:", font=("Arial", 10, "bold"), background=self.lb_background).grid(
            row=5, column=0, sticky="w", padx=10, pady=5)
        self.lbl_most_borrowed_book_val.grid(row=5, column=1, sticky="w", padx=10)

        tk. Label(self.frame_statics, text="Độc giả tích cực nhất:", font=("Arial", 10, "bold"), background=self.lb_background).grid(
            row=5, column=2, sticky="w", padx=10, pady=5)
        self.lbl_most_active_reader_val.grid(row=5, column=3, sticky="w", padx=10)

        # Phân cách
        ttk.Separator(self.frame_statics, orient='horizontal').grid(
            row=6, column=0, columnspan=4, sticky="ew", padx=10, pady=15)

        # Sách chưa trả
        tk.Label(self.frame_statics, text="SÁCH CHƯA TRẢ", font=("Arial", 12, "bold"), background=self.lb_background).grid(
            row=7, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))
        self.lbl_books_not_returned_count.grid(row=8, column=0, columnspan=4, sticky="w", padx=10)

        # Frame cho Treeview với scrollbar
        frame_not_returned = ttk.Frame(self.frame_statics)
        frame_not_returned.grid(row=9, column=0, columnspan=4, padx=10, pady=(5, 10), sticky="ew")
        
        scrollbar_not_returned = ttk.Scrollbar(frame_not_returned, orient="vertical")
        self.tree_books_not_returned = Treeview(
            frame_not_returned, 
            columns=("Sách", "Người mượn", "Ngày mượn", "Ngày trả dự kiến"), 
            show="headings", 
            height=5,
            yscrollcommand=scrollbar_not_returned.set
        )
        scrollbar_not_returned.config(command=self.tree_books_not_returned.yview)
        scrollbar_not_returned.pack(side="right", fill="y")
        self.tree_books_not_returned.pack(side="left", fill="both", expand=True)
        
        for col in ("Sách", "Người mượn", "Ngày mượn", "Ngày trả dự kiến"):
            self.tree_books_not_returned.heading(col, text=col)
            self.tree_books_not_returned.column(col, anchor="center", width=200)

        # Sách đến hẹn chưa trả
        tk.Label(self.frame_statics, text="SÁCH QUÁ HẠN", font=("Arial", 12, "bold"), background=self.lb_background).grid(
            row=10, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))
        self.lbl_books_due_not_returned_count.grid(row=11, column=0, columnspan=4, sticky="w", padx=10)

        frame_due_not_returned = ttk.Frame(self.frame_statics)
        frame_due_not_returned.grid(row=12, column=0, columnspan=4, padx=10, pady=(5, 10), sticky="ew")
        
        scrollbar_due_not_returned = ttk.Scrollbar(frame_due_not_returned, orient="vertical")
        self.tree_books_due_not_returned = Treeview(
            frame_due_not_returned,
            columns=("Sách", "Người mượn", "Ngày mượn", "Ngày trả dự kiến"), 
            show="headings", 
            height=5,
            yscrollcommand=scrollbar_due_not_returned. set
        )
        scrollbar_due_not_returned.config(command=self.tree_books_due_not_returned.yview)
        scrollbar_due_not_returned.pack(side="right", fill="y")
        self.tree_books_due_not_returned.pack(side="left", fill="both", expand=True)
        
        for col in ("Sách", "Người mượn", "Ngày mượn", "Ngày trả dự kiến"):
            self.tree_books_due_not_returned.heading(col, text=col)
            self.tree_books_due_not_returned.column(col, anchor="center", width=200)

        # Chi tiết số lượng sách
        tk.Label(self.frame_statics, text="CHI TIẾT SỐ LƯỢNG SÁCH", font=("Arial", 12, "bold"), background=self.lb_background).grid(
            row=13, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))

        frame_details = ttk.Frame(self. frame_statics)
        frame_details.grid(row=14, column=0, columnspan=4, padx=10, pady=(5, 10), sticky="ew")
        
        scrollbar_details = ttk. Scrollbar(frame_details, orient="vertical")
        self.tree_books_details = Treeview(
            frame_details,
            columns=("Sách", "Tổng số lượng", "Đang mượn", "Còn lại"), 
            show="headings", 
            height=8,
            yscrollcommand=scrollbar_details.set
        )
        scrollbar_details.config(command=self.tree_books_details.yview)
        scrollbar_details.pack(side="right", fill="y")
        self.tree_books_details.pack(side="left", fill="both", expand=True)
        
        for col in ("Sách", "Tổng số lượng", "Đang mượn", "Còn lại"):
            self.tree_books_details. heading(col, text=col)
            self.tree_books_details.column(col, anchor="center", width=200)

        # Nút làm mới và Export
        button_frame = ttk.Frame(self.frame_statics)
        button_frame.grid(row=15, column=0, columnspan=4, pady=20)
        
        btn_refresh = ttk.Button(button_frame, text="Làm mới dữ liệu", command=self.refresh_data)
        btn_refresh.pack(side="left", padx=5)
        
        btn_export = ttk.Button(button_frame, text="Xuất báo cáo Excel", command=self.export_to_excel)
        btn_export.pack(side="left", padx=5)

    def refresh_data(self):
        self.get_statistics()
        self.get_books_details()
        messagebox.showinfo("Thành công", "Đã làm mới dữ liệu!")

    def get_statistics(self):
        try:
            # Debug: In ra để kiểm tra
            
            # Tổng số lượng sách (tổng số lượng của tất cả các loại sách)
            self.cursor.execute("SELECT IFNULL(SUM(so_luong), 0) FROM Sach")
            total_books = self.cursor.fetchone()[0]
            self.lbl_total_books. config(text=str(total_books))

            # Tổng số độc giả
            self.cursor.execute("SELECT COUNT(*) FROM Doc_gia")
            total_readers = self.cursor.fetchone()[0]
            self.lbl_total_readers.config(text=str(total_readers))

            # Số sách đang được mượn (trang_thai = 0 nghĩa là chưa trả)
            self.cursor. execute("""
                SELECT COUNT(*) 
                FROM Muon_tra 
                WHERE trang_thai = 0
            """)
            borrowed_count = self.cursor.fetchone()[0]
            self.lbl_borrowed_books.config(text=str(borrowed_count))
            remaining_books = total_books - borrowed_count
            self.lbl_remaining_books.config(text=str(remaining_books))

            # Tỉ lệ trả sách đúng hạn
            # Sách đã trả (trang_thai = 1) VÀ trả đúng hạn (ngay_tra <= ngay_tra_du_kien)
            self.cursor. execute("""
                SELECT COUNT(*) 
                FROM Muon_tra 
                WHERE trang_thai = 1 
                AND ngay_tra <= ngay_tra_du_kien
            """)
            on_time_returns = self.cursor.fetchone()[0]
            
            # Tổng số lần đã trả
            self.cursor.execute("""
                SELECT COUNT(*) 
                FROM Muon_tra 
                WHERE trang_thai = 1
            """)
            total_returns = self.cursor.fetchone()[0]
            
            return_rate = (on_time_returns / total_returns * 100) if total_returns > 0 else 0
            self.lbl_return_rate.config(text=f"{return_rate:.2f}%")
            print(f"Tỉ lệ trả đúng hạn: {return_rate:.2f}% ({on_time_returns}/{total_returns})")

            # SÁCH CHƯA TRẢ (TẤT CẢ sách đang mượn - trang_thai = 0)
            self.cursor.execute("""
                SELECT 
                    Sach.ten_sach, 
                    Doc_gia.ten_doc_gia, 
                    DATE_FORMAT(Muon_tra.ngay_muon, '%Y-%m-%d') as ngay_muon,
                    DATE_FORMAT(Muon_tra.ngay_tra_du_kien, '%Y-%m-%d') as ngay_tra_du_kien
                FROM Muon_tra
                INNER JOIN Sach ON Muon_tra.ma_sach = Sach. ma_sach
                INNER JOIN Doc_gia ON Muon_tra.ma_doc_gia = Doc_gia. ma_doc_gia
                WHERE Muon_tra.trang_thai = 0
                ORDER BY Muon_tra. ngay_tra_du_kien
            """)
            books_not_returned = self.cursor.fetchall()
            self.update_treeview(self.tree_books_not_returned, books_not_returned)
            self.lbl_books_not_returned_count.config(text=f"Số lượng: {len(books_not_returned)}")

            # SÁCH ĐẾN HẠN MÀ CHƯA TRẢ (QUÁ HẠN)
            # trang_thai = 0 (chưa trả) VÀ ngay_tra_du_kien < ngày hiện tại
            self.cursor.execute("""
                SELECT 
                    Sach.ten_sach, 
                    Doc_gia.ten_doc_gia, 
                    DATE_FORMAT(Muon_tra.ngay_muon, '%Y-%m-%d') as ngay_muon,
                    DATE_FORMAT(Muon_tra.ngay_tra_du_kien, '%Y-%m-%d') as ngay_tra_du_kien
                FROM Muon_tra
                INNER JOIN Sach ON Muon_tra.ma_sach = Sach.ma_sach
                INNER JOIN Doc_gia ON Muon_tra.ma_doc_gia = Doc_gia.ma_doc_gia
                WHERE Muon_tra.trang_thai = 0 
                AND Muon_tra.ngay_tra_du_kien < CURDATE()
                ORDER BY Muon_tra. ngay_tra_du_kien
            """)
            books_due_not_returned = self.cursor.fetchall()
            print(f"Số lượng sách quá hạn: {len(books_due_not_returned)}")
            self. update_treeview(self. tree_books_due_not_returned, books_due_not_returned)
            self.lbl_books_due_not_returned_count.config(text=f"Số lượng: {len(books_due_not_returned)}")

            # Sách được mượn nhiều nhất
            self.cursor.execute("""
                SELECT Sach.ten_sach, COUNT(Muon_tra. ma_sach) as times
                FROM Sach
                INNER JOIN Muon_tra ON Sach. ma_sach = Muon_tra.ma_sach
                GROUP BY Sach.ma_sach, Sach.ten_sach
                ORDER BY times DESC
                LIMIT 1
            """)
            most_borrowed_book = self.cursor.fetchone()
            if most_borrowed_book:
                self.lbl_most_borrowed_book_val.config(text=f"{most_borrowed_book[0]} ({most_borrowed_book[1]} lần)")
            else:
                self.lbl_most_borrowed_book_val.config(text="Không có dữ liệu")

            # Độc giả tích cực nhất
            self.cursor.execute("""
                SELECT Doc_gia.ten_doc_gia, COUNT(Muon_tra.ma_doc_gia) as times
                FROM Doc_gia
                INNER JOIN Muon_tra ON Doc_gia.ma_doc_gia = Muon_tra. ma_doc_gia
                GROUP BY Doc_gia.ma_doc_gia, Doc_gia.ten_doc_gia
                ORDER BY times DESC
                LIMIT 1
            """)
            most_active_reader = self.cursor.fetchone()
            if most_active_reader:
                self.lbl_most_active_reader_val.config(text=f"{most_active_reader[0]} ({most_active_reader[1]} lần)")
            else:
                self.lbl_most_active_reader_val.config(text="Không có dữ liệu")
        except Exception as e:
            print(f"LỖI: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Lỗi truy vấn!", f"Lỗi khi thực hiện truy vấn:\n{e}")

    def get_books_details(self):
        try:            
            # Lấy thông tin chi tiết từng loại sách
            self.cursor.execute("""
                SELECT 
                    Sach.ten_sach, 
                    Sach.so_luong as total,
                    IFNULL((SELECT COUNT(*) 
                     FROM Muon_tra 
                     WHERE Muon_tra.ma_sach = Sach.ma_sach 
                     AND Muon_tra.trang_thai = 0), 0) AS borrowed
                FROM Sach
                ORDER BY Sach.ten_sach
            """)
            books_details = self.cursor.fetchall()
            
            print(f"Số loại sách: {len(books_details)}")
            
            # Xóa dữ liệu cũ
            self.update_treeview(self.tree_books_details, [])
            
            # Thêm dữ liệu mới
            for book, total, borrowed in books_details:
                borrowed = borrowed if borrowed is not None else 0
                remaining = total - borrowed
                
                print(f"{book}: Tổng={total}, Đang mượn={borrowed}, Còn lại={remaining}")
                
                self.tree_books_details.insert("", tk.END, values=(book, total, borrowed, remaining))            
        except Exception as e:
            print(f"LỖI: {e}")
            import traceback
            traceback. print_exc()
            messagebox.showerror("Lỗi truy vấn!", f"Lỗi khi lấy chi tiết sách:\n{e}")

    def update_treeview(self, treeview, data):
        """Cập nhật dữ liệu vào treeview"""
        # Xóa tất cả dữ liệu cũ
        for item in treeview.get_children():
            treeview. delete(item)
        
        # Thêm dữ liệu mới
        for row in data:
            treeview. insert("", tk.END, values=row)

    def export_to_excel(self):
        file_path = asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"ThongKeThuVien_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path:
            return
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook. active
            sheet.title = "Thống kê"

            # Tiêu đề
            sheet. merge_cells('A1:D1')
            title_cell = sheet['A1']
            title_cell.value = "BÁO CÁO THỐNG KÊ THƯ VIỆN"
            title_cell.font = openpyxl.styles.Font(size=16, bold=True)
            title_cell.alignment = openpyxl.styles. Alignment(horizontal='center')

            # Ngày xuất báo cáo
            sheet['A2'] = f"Ngày xuất báo cáo: {datetime. now().strftime('%d/%m/%Y %H:%M:%S')}"
            sheet. merge_cells('A2:D2')

            # Thống kê tổng quan
            row = 4
            sheet[f'A{row}'] = "I.  THỐNG KÊ TỔNG QUAN"
            sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12)
            row += 1
            
            stats = [
                ("Tổng số lượng sách:", self.lbl_total_books.cget("text")),
                ("Tổng số độc giả:", self.lbl_total_readers.cget("text")),
                ("Số sách đang được mượn:", self.lbl_borrowed_books.cget("text")),
                ("Số sách còn lại:", self.lbl_remaining_books. cget("text")),
                ("Tỉ lệ trả sách đúng hạn:", self.lbl_return_rate. cget("text")),
                ("Sách được mượn nhiều nhất:", self.lbl_most_borrowed_book_val.cget("text")),
                ("Độc giả tích cực nhất:", self.lbl_most_active_reader_val.cget("text")),
            ]
            
            for label, value in stats:
                sheet[f'A{row}'] = label
                sheet[f'B{row}'] = value
                sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                row += 1

            # Sách chưa trả
            row += 2
            sheet[f'A{row}'] = "II.  SÁCH CHƯA TRẢ"
            sheet[f'A{row}'].font = openpyxl.styles. Font(bold=True, size=12)
            row += 1
            sheet[f'A{row}'] = self.lbl_books_not_returned_count.cget("text")
            row += 1
            
            headers = ["Sách", "Người mượn", "Ngày mượn", "Ngày trả dự kiến"]
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row, column=col_idx, value=header)
                cell. font = openpyxl. styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            row += 1
            
            row = self.write_to_excel(sheet, self.tree_books_not_returned, start_row=row)

            # Sách quá hạn
            row += 2
            sheet[f'A{row}'] = "III.  SÁCH QUÁ HẠN"
            sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12)
            row += 1
            sheet[f'A{row}'] = self.lbl_books_due_not_returned_count.cget("text")
            row += 1
            
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row, column=col_idx, value=header)
                cell. font = openpyxl. styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            row += 1
            
            row = self.write_to_excel(sheet, self.tree_books_due_not_returned, start_row=row)

            # Chi tiết số lượng sách
            row += 2
            sheet[f'A{row}'] = "IV. CHI TIẾT SỐ LƯỢNG SÁCH"
            sheet[f'A{row}'].font = openpyxl.styles. Font(bold=True, size=12)
            row += 1
            
            detail_headers = ["Sách", "Tổng số lượng", "Đang mượn", "Còn lại"]
            for col_idx, header in enumerate(detail_headers, start=1):
                cell = sheet.cell(row=row, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            row += 1
            
            self.write_to_excel(sheet, self.tree_books_details, start_row=row)

            # Tự động điều chỉnh độ rộng cột
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell. value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(file_path)
            messagebox.showinfo("Thành công", f"Dữ liệu đã được xuất ra file:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất dữ liệu ra file Excel:\n{e}")

    def write_to_excel(self, sheet, treeview, start_row):
        """Ghi dữ liệu từ treeview vào Excel"""
        row_idx = start_row
        for item in treeview.get_children():
            values = treeview.item(item, "values")
            for col_idx, value in enumerate(values, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1
        return row_idx

    def tro_ve(self):
        """Quay về màn hình chính"""
        try:
            self.cursor.close()
            self.conn.close()
        except:
            pass
        self.frame_statics.destroy()
        from main import LibraryManagementScreen
        LibraryManagementScreen(self.root)

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1200x800") 
    app = LibraryStatisticsScreen(root)
    root.mainloop()